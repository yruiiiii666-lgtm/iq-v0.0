from __future__ import annotations

import json
import math
import re
import sqlite3
from dataclasses import asdict, dataclass
from datetime import datetime
from fractions import Fraction
from pathlib import Path
from typing import Callable, Iterable

import numpy as np
from matplotlib.figure import Figure
from scipy.signal import firwin, lfilter, resample_poly

from iq_reader import IQRecording, read_iq_contiguous, recording_from_paths
from scene_catalog import discover_iq_groups_by_prefix, list_linked_iq_details, list_scene_locations
from spectrum_feature_library import ALL_BAND, list_feature_records


MAX_RECONSTRUCTION_SAMPLES = 8_000_000
MAX_MEASURED_PLAYBACK_SAMPLES = 40_000_000
POWER_SCAN_BLOCK_SAMPLES = 8_000_000
REPRESENTATIVE_CACHE_VERSION = 1
REPRESENTATIVE_DEFAULTS = {
    "frequency_tolerance_mhz": 20.0,
    "sample_window_ms": 10.0,
    "polarization": "垂直极化",
    "band": ALL_BAND,
    "playback_segment_ms": 1000.0,
    "minimum_location_count": 2,
    "maximum_bands": 5,
}


@dataclass(frozen=True)
class IQCandidate:
    city: str
    point: str
    recording_stem: str
    center_frequency_mhz: float
    sample_rate_hz: float
    coverage_low_mhz: float
    coverage_high_mhz: float
    centrality_score: float
    available: bool
    wsm_file: str
    ws1_file: str
    ws2_file: str

    def recording(self) -> IQRecording:
        return recording_from_paths(
            self.recording_stem,
            Path(self.wsm_file),
            (Path(self.ws1_file), Path(self.ws2_file)),
        )


@dataclass(frozen=True)
class TypicalSignal:
    rank: int
    scene_type: str
    typical_frequency_mhz: float
    occurrence_count: int
    scene_location_count: int
    scene_probability: float
    global_probability: float
    specificity: float
    lift: float
    mean_level_dbuv_m: float
    global_median_level_dbuv_m: float
    level_contrast_db: float
    mean_bandwidth_3db_mhz: float
    category: str
    score: float
    location_keys: tuple[tuple[str, str], ...]
    iq_candidates: tuple[IQCandidate, ...]


@dataclass(frozen=True)
class TypicalSignalResult:
    scene_type: str
    polarization: str
    band: str
    tolerance_mhz: float
    minimum_probability: float
    scene_location_count: int
    global_location_count: int
    signals: tuple[TypicalSignal, ...]


@dataclass(frozen=True)
class SpectrumIQCorrespondence:
    serial: int
    city: str
    point: str
    scene_type: str
    polarization: str
    band: str
    spectrum_peak_rank: int | None
    spectrum_frequency_mhz: float | None
    spectrum_level_dbuv_m: float | None
    spectrum_bandwidth_3db_mhz: float | None
    iq_center_frequency_mhz: float | None
    frequency_difference_mhz: float | None
    iq_coverage_low_mhz: float | None
    iq_coverage_high_mhz: float | None
    recording_stem: str
    iq_available: bool
    match_status: str


@dataclass(frozen=True)
class LocationSpectrumIQSummary:
    serial: int
    city: str
    point: str
    scene_type: str
    polarization: str
    band: str
    spectrum_peak_frequencies_mhz: str
    spectrum_peak_levels_dbuv_m: str
    spectrum_bandwidths_3db_mhz: str
    iq_center_frequencies_mhz: str
    iq_recording_stems: str
    iq_file_statuses: str
    correspondence_summary: str


@dataclass(frozen=True)
class SceneIQRepresentative:
    rank: int
    scene_type: str
    group_center_frequency_mhz: float
    group_low_frequency_mhz: float
    group_high_frequency_mhz: float
    candidate_count: int
    location_count: int
    scene_location_count: int
    representative_frequency_mhz: float
    representative_score: float
    spectrum_support_count: int
    spectrum_location_count: int
    spectrum_median_level_dbuv_m: float | None
    spectrum_median_bandwidth_mhz: float | None
    city: str
    point: str
    recording_stem: str
    center_frequency_mhz: float
    relative_power_dbfs: float
    estimated_power_dbm: float
    reference_level_dbm: float
    selected_start_s: float
    selected_end_s: float
    selected_duration_s: float
    detected_event_duration_s: float
    recording_duration_s: float
    sample_window_ms: float
    scanned_window_count: int
    rejected_clipped_windows: int
    selection_note: str
    sample_rate_hz: float
    wsm_file: str
    ws1_file: str
    ws2_file: str


def _representative_cache_connection(database_path: Path) -> sqlite3.Connection:
    database_path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(database_path)
    connection.execute("PRAGMA journal_mode = WAL")
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS scene_iq_representative_cache (
            scene_type TEXT PRIMARY KEY,
            cache_version INTEGER NOT NULL,
            parameters_json TEXT NOT NULL,
            results_json TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    return connection


def load_scene_iq_representative_cache(
    database_path: Path,
    scene_type: str,
) -> tuple[tuple[SceneIQRepresentative, ...], str] | None:
    """Return one scene's cached representative IQ rows, if generated by this algorithm version."""
    connection = _representative_cache_connection(database_path)
    try:
        row = connection.execute(
            "SELECT cache_version, results_json, updated_at "
            "FROM scene_iq_representative_cache WHERE scene_type=?",
            (scene_type,),
        ).fetchone()
    finally:
        connection.close()
    if row is None or int(row[0]) != REPRESENTATIVE_CACHE_VERSION:
        return None
    try:
        payload = json.loads(str(row[1]))
        results = tuple(SceneIQRepresentative(**item) for item in payload)
    except (TypeError, ValueError, json.JSONDecodeError):
        return None
    return results, str(row[2])


def save_scene_iq_representative_cache(
    database_path: Path,
    scene_type: str,
    rows: tuple[SceneIQRepresentative, ...],
    parameters: dict[str, object] | None = None,
) -> str:
    updated_at = datetime.now().isoformat(timespec="seconds")
    connection = _representative_cache_connection(database_path)
    try:
        connection.execute(
            """
            INSERT INTO scene_iq_representative_cache
                (scene_type, cache_version, parameters_json, results_json, updated_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(scene_type) DO UPDATE SET
                cache_version=excluded.cache_version,
                parameters_json=excluded.parameters_json,
                results_json=excluded.results_json,
                updated_at=excluded.updated_at
            """,
            (
                scene_type,
                REPRESENTATIVE_CACHE_VERSION,
                json.dumps(parameters or REPRESENTATIVE_DEFAULTS, ensure_ascii=False, sort_keys=True),
                json.dumps([asdict(item) for item in rows], ensure_ascii=False),
                updated_at,
            ),
        )
        connection.commit()
    finally:
        connection.close()
    return updated_at


def export_scene_iq_representative_cache(database_path: Path, output_path: Path) -> Path:
    """Export every cached scene to a ready-to-open Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    connection = _representative_cache_connection(database_path)
    try:
        cached = connection.execute(
            "SELECT scene_type, parameters_json, results_json, updated_at "
            "FROM scene_iq_representative_cache ORDER BY scene_type"
        ).fetchall()
    finally:
        connection.close()
    if not cached:
        raise ValueError("数据库中还没有场景代表IQ筛选缓存")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "缓存说明"
    summary.append(("场景类型", "缓存时间", "代表频段数", "固定筛选规则"))
    headers = (
        "排序", "代表频段(MHz)", "频谱参考频率(MHz)", "IQ地点覆盖", "频谱地点支持",
        "代表得分", "城市", "采集地点", "IQ数据组", "中心频率(MHz)",
        "估算功率(dBm)", "相对功率(dBFS)", "检测事件(ms)", "回放片段起点(s)",
        "回放片段终点(s)", "回放片段长度(s)", "原记录时长(s)", "采样率(MS/s)",
        "选择说明", "WSM文件", "WS1文件", "WS2文件",
    )
    used_names = {summary.title}
    for scene_type, parameters_json, results_json, updated_at in cached:
        payload = json.loads(results_json)
        parameters = json.loads(parameters_json)
        summary.append((scene_type, updated_at, len(payload), json.dumps(parameters, ensure_ascii=False)))
        base_name = re.sub(r"[\\/*?:\[\]]", "_", str(scene_type))[:31] or "未命名场景"
        sheet_name = base_name
        suffix = 2
        while sheet_name in used_names:
            tail = f"_{suffix}"
            sheet_name = f"{base_name[:31 - len(tail)]}{tail}"
            suffix += 1
        used_names.add(sheet_name)
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for item in payload:
            low = float(item["group_low_frequency_mhz"])
            high = float(item["group_high_frequency_mhz"])
            frequency_group = f"{low:g}" if math.isclose(low, high) else f"{low:g}～{high:g}"
            sheet.append((
                item["rank"], frequency_group, item["representative_frequency_mhz"],
                f"{item['location_count']}/{item['scene_location_count']}",
                f"{item['spectrum_support_count']}/{item['spectrum_location_count']}",
                item["representative_score"], item["city"], item["point"], item["recording_stem"],
                item["center_frequency_mhz"], item["estimated_power_dbm"], item["relative_power_dbfs"],
                float(item["detected_event_duration_s"]) * 1e3, item["selected_start_s"],
                item["selected_end_s"], item["selected_duration_s"], item["recording_duration_s"],
                float(item["sample_rate_hz"]) / 1e6, item["selection_note"], item["wsm_file"],
                item["ws1_file"], item["ws2_file"],
            ))
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D4ED8")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in sheet.columns:
            values = [len(str(cell.value or "")) for cell in column[:80]]
            sheet.column_dimensions[column[0].column_letter].width = min(max(max(values, default=8) + 2, 10), 42)
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D4ED8")
    summary.freeze_panes = "A2"
    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 22
    summary.column_dimensions["C"].width = 14
    summary.column_dimensions["D"].width = 70
    workbook.save(output_path)
    return output_path


@dataclass(frozen=True)
class _PowerScanResult:
    relative_power_dbfs: float
    estimated_power_dbm: float
    selected_start_s: float
    selected_end_s: float
    detected_event_duration_s: float
    scanned_window_count: int
    rejected_clipped_windows: int
    selection_note: str


@dataclass(frozen=True)
class ReconstructionComponent:
    name: str
    source_type: str
    modulation: str
    frequency_mhz: float
    offset_mhz: float
    relative_level_db: float
    bandwidth_mhz: float
    source_reference: str = ""


@dataclass
class ReconstructionResult:
    name: str
    mode: str
    iq: np.ndarray
    sample_rate_hz: float
    center_frequency_mhz: float
    components: tuple[ReconstructionComponent, ...]
    metadata: dict[str, object]
    figure: Figure
    original_iq: np.ndarray | None = None
    original_sample_rate_hz: float | None = None
    original_center_frequency_mhz: float | None = None

    @property
    def duration_s(self) -> float:
        return self.iq.size / self.sample_rate_hz


def _safe_json_list(value: str) -> list[dict[str, float | int]]:
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, list) else []
    except (TypeError, ValueError):
        return []


def _cluster_peak_observations(
    observations: list[dict[str, object]], tolerance_mhz: float
) -> list[dict[tuple[str, str], dict[str, object]]]:
    observations.sort(key=lambda item: float(item["frequency"]))
    clusters: list[dict[tuple[str, str], dict[str, object]]] = []
    for observation in observations:
        frequency = float(observation["frequency"])
        candidates: list[tuple[float, int]] = []
        for index, cluster in enumerate(clusters):
            center = float(np.mean([float(item["frequency"]) for item in cluster.values()]))
            distance = abs(frequency - center)
            if distance <= tolerance_mhz:
                candidates.append((distance, index))
        location_key = observation["location_key"]
        assert isinstance(location_key, tuple)
        if candidates:
            cluster = clusters[min(candidates)[1]]
            previous = cluster.get(location_key)
            if previous is None or float(observation["level"]) > float(previous["level"]):
                cluster[location_key] = observation
        else:
            clusters.append({location_key: observation})
    return clusters


def _candidate_iq_links(
    database_path: Path,
    location_keys: Iterable[tuple[str, str]],
    frequency_mhz: float,
    guard_mhz: float = 1.0,
) -> tuple[IQCandidate, ...]:
    candidates: list[IQCandidate] = []
    seen: set[tuple[str, str, str]] = set()
    for city, point in location_keys:
        for link in list_linked_iq_details(database_path, city, point):
            key = (city, point, link.recording_stem)
            if key in seen:
                continue
            seen.add(key)
            available = True
            try:
                recording = recording_from_paths(
                    link.recording_stem,
                    Path(link.wsm_file),
                    (Path(link.ws1_file), Path(link.ws2_file)),
                )
                center_frequency_mhz = recording.center_frequency_mhz
                sample_rate_hz = recording.sample_rate_hz
            except (FileNotFoundError, OSError, ValueError):
                available = False
                matches = re.findall(r"(\d+(?:\.\d+)?)", link.recording_stem)
                if not matches:
                    continue
                center_frequency_mhz = float(matches[-1])
                sample_rate_hz = 40e6
            half_band_mhz = sample_rate_hz / 2e6
            usable_half_band = max(0.0, half_band_mhz - guard_mhz)
            distance = abs(frequency_mhz - center_frequency_mhz)
            if distance > usable_half_band:
                continue
            centrality = 1.0 - distance / max(usable_half_band, 1e-9)
            candidates.append(IQCandidate(
                city=city,
                point=point,
                recording_stem=link.recording_stem,
                center_frequency_mhz=center_frequency_mhz,
                sample_rate_hz=sample_rate_hz,
                coverage_low_mhz=center_frequency_mhz - half_band_mhz,
                coverage_high_mhz=center_frequency_mhz + half_band_mhz,
                centrality_score=centrality,
                available=available,
                wsm_file=link.wsm_file,
                ws1_file=link.ws1_file,
                ws2_file=link.ws2_file,
            ))
    candidates.sort(key=lambda item: (not item.available, -item.centrality_score, item.city, item.point, item.recording_stem))
    return tuple(candidates)


def _linked_iq_for_location(
    database_path: Path,
    city: str,
    point: str,
    iq_root: Path | None = None,
    relative_directory: str = "",
    recording_prefix: str = "",
) -> tuple[IQCandidate, ...]:
    candidates: list[IQCandidate] = []
    def belongs_to_prefix(stem: str) -> bool:
        if not recording_prefix:
            return True
        if not stem.casefold().startswith(recording_prefix.casefold()):
            return False
        return re.fullmatch(r"\d+(?:\.\d+)?m?", stem[len(recording_prefix):], re.IGNORECASE) is not None

    links_by_stem = {
        link.recording_stem: link
        for link in list_linked_iq_details(database_path, city, point)
        if belongs_to_prefix(link.recording_stem)
    }
    if iq_root is not None and recording_prefix:
        for link in discover_iq_groups_by_prefix(iq_root, relative_directory, recording_prefix):
            links_by_stem[link.recording_stem] = link
    for link in links_by_stem.values():
        available = True
        matches = re.findall(r"(\d+(?:\.\d+)?)", link.recording_stem)
        filename_center = float(matches[-1]) if matches else None
        try:
            recording = recording_from_paths(
                link.recording_stem,
                Path(link.wsm_file),
                (Path(link.ws1_file), Path(link.ws2_file)),
            )
            center = filename_center if filename_center is not None else recording.center_frequency_mhz
            sample_rate = recording.sample_rate_hz
        except (FileNotFoundError, OSError, ValueError):
            available = False
            if filename_center is None:
                continue
            center = filename_center
            sample_rate = 40e6
        half_band = sample_rate / 2e6
        candidates.append(IQCandidate(
            city=city,
            point=point,
            recording_stem=link.recording_stem,
            center_frequency_mhz=center,
            sample_rate_hz=sample_rate,
            coverage_low_mhz=center - half_band,
            coverage_high_mhz=center + half_band,
            centrality_score=1.0,
            available=available,
            wsm_file=link.wsm_file,
            ws1_file=link.ws1_file,
            ws2_file=link.ws2_file,
        ))
    candidates.sort(key=lambda item: (item.center_frequency_mhz, item.recording_stem.casefold()))
    return tuple(candidates)


def build_spectrum_iq_correspondence(
    database_path: Path,
    polarization: str = "垂直极化",
    band: str = ALL_BAND,
    scene_type: str = "全部",
    guard_mhz: float = 1.0,
    iq_root: Path | None = None,
) -> tuple[SpectrumIQCorrespondence, ...]:
    """Build a lossless two-way table between spectrum peaks and linked IQ recordings."""
    if guard_mhz < 0:
        raise ValueError("IQ带宽保护量不能为负数")
    locations = list_scene_locations(database_path, scene_type=scene_type)
    records = list_feature_records(database_path, polarization=polarization, band=band)
    record_map = {(record.city, record.point): record for record in records}
    rows: list[SpectrumIQCorrespondence] = []
    serial = 0

    def append_row(
        location,
        peak: dict[str, float | int] | None,
        iq: IQCandidate | None,
        status: str,
    ) -> None:
        nonlocal serial
        serial += 1
        spectrum_frequency = float(peak.get("frequency_mhz", 0.0)) if peak is not None else None
        iq_center = iq.center_frequency_mhz if iq is not None else None
        rows.append(SpectrumIQCorrespondence(
            serial=serial,
            city=location.city,
            point=location.point,
            scene_type=location.scene_type,
            polarization=polarization,
            band=band,
            spectrum_peak_rank=int(peak.get("rank", 0)) if peak is not None else None,
            spectrum_frequency_mhz=spectrum_frequency,
            spectrum_level_dbuv_m=float(peak.get("field_dbuv_m", 0.0)) if peak is not None else None,
            spectrum_bandwidth_3db_mhz=float(peak.get("bandwidth_3db_mhz", 0.0)) if peak is not None else None,
            iq_center_frequency_mhz=iq_center,
            frequency_difference_mhz=(abs(spectrum_frequency - iq_center) if spectrum_frequency is not None and iq_center is not None else None),
            iq_coverage_low_mhz=iq.coverage_low_mhz if iq is not None else None,
            iq_coverage_high_mhz=iq.coverage_high_mhz if iq is not None else None,
            recording_stem=iq.recording_stem if iq is not None else "",
            iq_available=iq.available if iq is not None else False,
            match_status=status,
        ))

    for location in locations:
        record = record_map.get((location.city, location.spectrum_point or location.point))
        peaks = _safe_json_list(record.top_peaks_json) if record is not None else []
        iq_candidates = _linked_iq_for_location(
            database_path,
            location.city,
            location.point,
            iq_root=iq_root,
            relative_directory=location.iq_relative_directory,
            recording_prefix=location.iq_recording_prefix,
        )
        matched_iq: set[str] = set()
        if peaks:
            for peak in peaks:
                frequency = float(peak.get("frequency_mhz", 0.0))
                covering = [
                    candidate for candidate in iq_candidates
                    if candidate.coverage_low_mhz + guard_mhz <= frequency <= candidate.coverage_high_mhz - guard_mhz
                ]
                if not covering:
                    append_row(location, peak, None, "频谱峰值无IQ覆盖")
                    continue
                covering.sort(key=lambda item: (abs(frequency - item.center_frequency_mhz), not item.available))
                for candidate in covering:
                    matched_iq.add(candidate.recording_stem)
                    difference = abs(frequency - candidate.center_frequency_mhz)
                    peak_bandwidth = float(peak.get("bandwidth_3db_mhz", 0.0))
                    close_limit = max(2.0, peak_bandwidth / 2.0)
                    status = "中心频率匹配" if difference <= close_limit else "IQ带宽覆盖"
                    append_row(location, peak, candidate, status)
        for candidate in iq_candidates:
            if candidate.recording_stem not in matched_iq:
                status = "仅有IQ，无频谱数据" if record is None else "IQ中心频率无主要峰值"
                append_row(location, None, candidate, status)
        if not peaks and not iq_candidates:
            append_row(location, None, None, "无频谱特征且无关联IQ")
        elif record is not None and not peaks and iq_candidates:
            # IQ-only rows above already describe the missing peak condition.
            pass
        elif record is not None and peaks and not iq_candidates:
            # Peak-only rows above already describe the missing IQ condition.
            pass
    return tuple(rows)


def aggregate_spectrum_iq_correspondence(
    rows: Iterable[SpectrumIQCorrespondence],
) -> tuple[LocationSpectrumIQSummary, ...]:
    """Collapse detailed matching rows into exactly one CSV-friendly row per location."""
    grouped: dict[tuple[str, str], list[SpectrumIQCorrespondence]] = {}
    for row in rows:
        grouped.setdefault((row.city, row.point), []).append(row)

    summaries: list[LocationSpectrumIQSummary] = []
    for serial, location_rows in enumerate(grouped.values(), start=1):
        first = location_rows[0]
        peaks: dict[tuple[int, float], tuple[float, float, float]] = {}
        iq_groups: dict[str, tuple[float, bool]] = {}
        matched_peak_keys: set[tuple[int, float]] = set()
        for row in location_rows:
            if row.spectrum_peak_rank is not None and row.spectrum_frequency_mhz is not None:
                key = (row.spectrum_peak_rank, row.spectrum_frequency_mhz)
                peaks[key] = (
                    row.spectrum_frequency_mhz,
                    row.spectrum_level_dbuv_m or 0.0,
                    row.spectrum_bandwidth_3db_mhz or 0.0,
                )
                if row.match_status in {"中心频率匹配", "IQ带宽覆盖"}:
                    matched_peak_keys.add(key)
            if row.recording_stem and row.iq_center_frequency_mhz is not None:
                iq_groups[row.recording_stem] = (row.iq_center_frequency_mhz, row.iq_available)

        ordered_peaks = [peaks[key] for key in sorted(peaks, key=lambda item: (item[0], item[1]))]
        ordered_iq = sorted(
            ((center, stem, available) for stem, (center, available) in iq_groups.items()),
            key=lambda item: (item[0], item[1].casefold()),
        )
        peak_count = len(ordered_peaks)
        iq_count = len(ordered_iq)
        if peak_count and iq_count:
            relation = f"已覆盖主要峰值{len(matched_peak_keys)}/{peak_count}，关联IQ {iq_count}组"
        elif peak_count:
            relation = f"有{peak_count}个主要峰值，无关联IQ"
        elif iq_count:
            relation = f"无频谱数据，仅有IQ {iq_count}组"
        else:
            relation = "无频谱数据，无关联IQ"
        summaries.append(LocationSpectrumIQSummary(
            serial=serial,
            city=first.city,
            point=first.point,
            scene_type=first.scene_type,
            polarization=first.polarization,
            band=first.band,
            spectrum_peak_frequencies_mhz=", ".join(f"{item[0]:.6f}" for item in ordered_peaks),
            spectrum_peak_levels_dbuv_m=", ".join(f"{item[1]:.3f}" for item in ordered_peaks),
            spectrum_bandwidths_3db_mhz=", ".join(f"{item[2]:.6f}" for item in ordered_peaks),
            iq_center_frequencies_mhz=", ".join(f"{item[0]:g}" for item in ordered_iq),
            iq_recording_stems=", ".join(item[1] for item in ordered_iq),
            iq_file_statuses=", ".join("可读取" if item[2] else "离线" for item in ordered_iq),
            correspondence_summary=relation,
        ))
    return tuple(summaries)


def _scan_recording_power(
    recording: IQRecording,
    sample_window_ms: float,
    playback_segment_ms: float = 1000.0,
) -> _PowerScanResult:
    """Locate the strongest stable event, then retain a longer playback segment around it."""
    sample_rate = recording.sample_rate_hz
    if sample_rate <= 0 or recording.total_samples <= 0:
        raise ValueError(f"{recording.stem}没有有效采样率或采样点")
    window_samples = min(
        recording.total_samples,
        max(4096, int(round(sample_rate * sample_window_ms / 1000.0))),
    )
    window_count = max(1, recording.total_samples // window_samples)
    windows_per_block = max(1, min(32, POWER_SCAN_BLOCK_SAMPLES // window_samples))
    powers: list[np.ndarray] = []
    clipped_flags: list[np.ndarray] = []

    for first_window in range(0, window_count, windows_per_block):
        block_window_count = min(windows_per_block, window_count - first_window)
        block_start = first_window * window_samples
        iq = read_iq_contiguous(recording, block_start, block_window_count * window_samples)
        complete_windows = iq.size // window_samples
        if complete_windows <= 0:
            continue
        matrix = iq[: complete_windows * window_samples].reshape(complete_windows, window_samples)
        complex_mean = np.mean(matrix, axis=1, dtype=np.complex128)
        mean_square = np.mean(np.abs(matrix) ** 2, axis=1, dtype=np.float64)
        powers.append(np.maximum(mean_square - np.abs(complex_mean) ** 2, 1e-15))

        component_peak = np.maximum(np.abs(matrix.real), np.abs(matrix.imag))
        clipped_ratio = np.mean(component_peak >= 0.9995, axis=1)
        clipped_flags.append(clipped_ratio >= 0.001)

    if not powers:
        raise ValueError(f"{recording.stem}无法读取功率扫描窗口")
    linear_power = np.concatenate(powers)
    clipped = np.concatenate(clipped_flags)
    power_db = 10.0 * np.log10(np.maximum(linear_power, 1e-15))
    valid = ~clipped
    selection_note = "整条IQ稳定RMS功率扫描"

    stable_power = linear_power.copy()
    stable_valid = valid.copy()
    uses_stable_triplet = False
    if linear_power.size >= 3:
        stable_power = np.convolve(linear_power, np.ones(3, dtype=np.float64) / 3.0, mode="same")
        local_spread = np.full(linear_power.size, math.inf, dtype=np.float64)
        local_spread[1:-1] = (
            np.maximum.reduce((power_db[:-2], power_db[1:-1], power_db[2:]))
            - np.minimum.reduce((power_db[:-2], power_db[1:-1], power_db[2:]))
        )
        stable_valid[:] = False
        stable_valid[1:-1] = valid[:-2] & valid[1:-1] & valid[2:] & (local_spread[1:-1] <= 6.0)
        uses_stable_triplet = bool(np.any(stable_valid))
    if not np.any(stable_valid):
        stable_power = linear_power
        stable_valid = valid
        selection_note = "整条IQ RMS功率扫描（未找到连续3窗稳定事件，已回退到单窗最大值）"
    if not np.any(stable_valid):
        stable_valid = np.ones(linear_power.size, dtype=bool)
        selection_note = "整条IQ RMS功率扫描（所有窗口均疑似削顶，请复核源数据）"

    ranked_power = np.where(stable_valid, stable_power, -math.inf)
    best_index = int(np.argmax(ranked_power))
    best_power = 10.0 * math.log10(max(float(stable_power[best_index]), 1e-15))

    baseline = float(np.median(power_db[valid])) if np.any(valid) else float(np.median(power_db))
    event_threshold = max(baseline + 3.0, best_power - 6.0)
    event_left = best_index - 1 if uses_stable_triplet and best_index > 0 else best_index
    event_right = (
        best_index + 1
        if uses_stable_triplet and best_index + 1 < power_db.size
        else best_index
    )
    while event_left > 0 and valid[event_left - 1] and power_db[event_left - 1] >= event_threshold:
        event_left -= 1
    while (
        event_right + 1 < power_db.size
        and valid[event_right + 1]
        and power_db[event_right + 1] >= event_threshold
    ):
        event_right += 1

    event_start = event_left * window_samples
    event_end = min(recording.total_samples, (event_right + 1) * window_samples)
    detected_event_samples = max(window_samples, event_end - event_start)
    requested_segment_samples = max(
        window_samples,
        int(round(sample_rate * playback_segment_ms / 1000.0)),
    )
    selected_samples = min(
        recording.total_samples,
        requested_segment_samples,
        MAX_MEASURED_PLAYBACK_SAMPLES,
    )
    best_center = int(round((best_index + 0.5) * window_samples))
    selected_start = best_center - selected_samples // 2
    selected_start = max(0, min(selected_start, recording.total_samples - selected_samples))
    selected_end = selected_start + selected_samples
    selection_note += (
        f"；检测到{detected_event_samples / sample_rate * 1e3:.1f} ms稳定高功率事件，"
        f"围绕最强时刻截取{selected_samples / sample_rate:.3f} s真实IQ用于回放"
    )
    if selected_samples < requested_segment_samples:
        reason = "原记录较短" if recording.total_samples < requested_segment_samples else "单次缓冲区限制"
        selection_note += f"（受{reason}影响，目标播放更长时自动循环）"

    reference = recording.reference_level_dbm
    estimated = best_power + reference if math.isfinite(reference) else best_power
    return _PowerScanResult(
        relative_power_dbfs=best_power,
        estimated_power_dbm=estimated,
        selected_start_s=selected_start / sample_rate,
        selected_end_s=selected_end / sample_rate,
        detected_event_duration_s=detected_event_samples / sample_rate,
        scanned_window_count=int(linear_power.size),
        rejected_clipped_windows=int(np.count_nonzero(clipped)),
        selection_note=selection_note,
    )


def _scene_spectrum_observations(
    database_path: Path,
    locations: Iterable[object],
    polarization: str,
    band: str,
) -> tuple[list[dict[str, object]], int]:
    records = list_feature_records(database_path, polarization=polarization, band=band)
    record_map = {(record.city, record.point): record for record in records}
    observations: list[dict[str, object]] = []
    spectrum_locations: set[tuple[str, str]] = set()
    for location in locations:
        city = str(getattr(location, "city"))
        point = str(getattr(location, "point"))
        spectrum_point = str(getattr(location, "spectrum_point", "") or point)
        record = record_map.get((city, spectrum_point))
        if record is None:
            continue
        location_key = (city, point)
        peaks = _safe_json_list(record.top_peaks_json)
        if peaks:
            spectrum_locations.add(location_key)
        for peak in peaks:
            observations.append({
                "location_key": location_key,
                "frequency": float(peak.get("frequency_mhz", 0.0)),
                "level": float(peak.get("field_dbuv_m", 0.0)),
                "bandwidth": float(peak.get("bandwidth_3db_mhz", 0.0)),
            })
    return observations, len(spectrum_locations)


def select_scene_iq_representatives(
    database_path: Path,
    scene_type: str,
    frequency_tolerance_mhz: float = 20.0,
    sample_window_ms: float = 10.0,
    polarization: str = "垂直极化",
    band: str = ALL_BAND,
    playback_segment_ms: float = 1000.0,
    minimum_location_count: int = 2,
    maximum_bands: int = 5,
    progress: Callable[[int, int, str], None] | None = None,
) -> tuple[SceneIQRepresentative, ...]:
    """Rank representative bands and retain a useful playback segment around each strongest event."""
    if frequency_tolerance_mhz <= 0:
        raise ValueError("相似频率容差必须大于0 MHz")
    if sample_window_ms <= 0:
        raise ValueError("功率抽样窗口必须大于0 ms")
    if playback_segment_ms <= 0:
        raise ValueError("实际回放片段长度必须大于0 ms")
    if minimum_location_count <= 0 or maximum_bands <= 0:
        raise ValueError("最少地点支持和最多代表频段必须是大于0的整数")
    locations = list_scene_locations(database_path, scene_type=scene_type)
    pending: list[tuple[object, object]] = []
    iq_location_keys: set[tuple[str, str]] = set()
    for location in locations:
        links = list_linked_iq_details(database_path, location.city, location.point)
        if links:
            iq_location_keys.add((location.city, location.point))
        for link in links:
            pending.append((location, link))
    if not pending:
        raise ValueError(f"{scene_type}没有已关联的IQ数据")

    candidates: list[dict[str, object]] = []
    errors: list[str] = []
    total = len(pending)
    for index, (location, link) in enumerate(pending, start=1):
        try:
            recording = recording_from_paths(
                link.recording_stem,
                Path(link.wsm_file),
                (Path(link.ws1_file), Path(link.ws2_file)),
            )
            scan = _scan_recording_power(recording, sample_window_ms, playback_segment_ms)
            candidates.append({
                "location": location,
                "link": link,
                "recording": recording,
                "frequency": recording.center_frequency_mhz,
                "scan": scan,
            })
        except (FileNotFoundError, OSError, ValueError) as exc:
            errors.append(f"{link.recording_stem}: {exc}")
        if progress is not None:
            progress(index, total, link.recording_stem)
    if not candidates:
        detail = errors[0] if errors else "未知读取错误"
        raise ValueError(f"没有可用于功率比较的完整IQ数据：{detail}")

    candidates.sort(key=lambda item: float(item["frequency"]))
    clusters: list[list[dict[str, object]]] = []
    for candidate in candidates:
        frequency = float(candidate["frequency"])
        if not clusters:
            clusters.append([candidate])
            continue
        center = float(np.mean([float(item["frequency"]) for item in clusters[-1]]))
        if abs(frequency - center) <= frequency_tolerance_mhz:
            clusters[-1].append(candidate)
        else:
            clusters.append([candidate])

    spectrum_observations, spectrum_location_count = _scene_spectrum_observations(
        database_path, locations, polarization, band
    )
    spectrum_tolerance_mhz = min(2.0, max(0.1, frequency_tolerance_mhz * 0.1))
    ranked_clusters: list[dict[str, object]] = []
    for cluster in clusters:
        coverage_low = min(
            float(item["frequency"]) - float(item["recording"].sample_rate_hz) / 2e6
            for item in cluster
        )
        coverage_high = max(
            float(item["frequency"]) + float(item["recording"].sample_rate_hz) / 2e6
            for item in cluster
        )
        covered_peaks = [
            item for item in spectrum_observations
            if coverage_low <= float(item["frequency"]) <= coverage_high
        ]
        peak_clusters = _cluster_peak_observations(covered_peaks, spectrum_tolerance_mhz)
        best_peak_cluster = max(
            peak_clusters,
            key=lambda items: (
                len(items),
                float(np.median([float(item["level"]) for item in items.values()])),
            ),
            default={},
        )
        spectrum_support_count = len(best_peak_cluster)
        if best_peak_cluster:
            representative_frequency = float(np.median([
                float(item["frequency"]) for item in best_peak_cluster.values()
            ]))
            spectrum_median_level = float(np.median([
                float(item["level"]) for item in best_peak_cluster.values()
            ]))
            spectrum_median_bandwidth = float(np.median([
                float(item["bandwidth"]) for item in best_peak_cluster.values()
            ]))
        else:
            representative_frequency = float(np.mean([float(item["frequency"]) for item in cluster]))
            spectrum_median_level = None
            spectrum_median_bandwidth = None

        cluster_location_keys = {
            (str(item["location"].city), str(item["location"].point)) for item in cluster
        }
        iq_coverage = len(cluster_location_keys) / max(len(iq_location_keys), 1)
        if spectrum_location_count:
            spectrum_support = spectrum_support_count / spectrum_location_count
            representative_score = 0.65 * iq_coverage + 0.35 * spectrum_support
        else:
            representative_score = iq_coverage
        selected = max(
            cluster,
            key=lambda item: float(item["scan"].estimated_power_dbm),
        )
        ranked_clusters.append({
            "cluster": cluster,
            "selected": selected,
            "location_count": len(cluster_location_keys),
            "representative_frequency": representative_frequency,
            "representative_score": representative_score,
            "spectrum_support_count": spectrum_support_count,
            "spectrum_median_level": spectrum_median_level,
            "spectrum_median_bandwidth": spectrum_median_bandwidth,
        })

    ranked_clusters.sort(key=lambda item: (
        -float(item["representative_score"]),
        -int(item["location_count"]),
        float(item["representative_frequency"]),
    ))
    qualified_clusters = [
        item for item in ranked_clusters
        if int(item["location_count"]) >= minimum_location_count
        or int(item["spectrum_support_count"]) >= minimum_location_count
    ]
    if not qualified_clusters:
        qualified_clusters = ranked_clusters[:1]
    ranked_clusters = qualified_clusters[:maximum_bands]
    representatives: list[SceneIQRepresentative] = []
    for rank, ranked in enumerate(ranked_clusters, start=1):
        cluster = ranked["cluster"]
        selected = ranked["selected"]
        assert isinstance(cluster, list) and isinstance(selected, dict)
        location = selected["location"]
        link = selected["link"]
        recording = selected["recording"]
        scan = selected["scan"]
        assert isinstance(scan, _PowerScanResult)
        frequencies = [float(item["frequency"]) for item in cluster]
        representatives.append(SceneIQRepresentative(
            rank=rank,
            scene_type=scene_type,
            group_center_frequency_mhz=float(np.mean(frequencies)),
            group_low_frequency_mhz=min(frequencies),
            group_high_frequency_mhz=max(frequencies),
            candidate_count=len(cluster),
            location_count=int(ranked["location_count"]),
            scene_location_count=len(locations),
            representative_frequency_mhz=float(ranked["representative_frequency"]),
            representative_score=float(ranked["representative_score"]),
            spectrum_support_count=int(ranked["spectrum_support_count"]),
            spectrum_location_count=spectrum_location_count,
            spectrum_median_level_dbuv_m=ranked["spectrum_median_level"],
            spectrum_median_bandwidth_mhz=ranked["spectrum_median_bandwidth"],
            city=location.city,
            point=location.point,
            recording_stem=link.recording_stem,
            center_frequency_mhz=float(selected["frequency"]),
            relative_power_dbfs=scan.relative_power_dbfs,
            estimated_power_dbm=scan.estimated_power_dbm,
            reference_level_dbm=recording.reference_level_dbm,
            selected_start_s=scan.selected_start_s,
            selected_end_s=scan.selected_end_s,
            selected_duration_s=scan.selected_end_s - scan.selected_start_s,
            detected_event_duration_s=scan.detected_event_duration_s,
            recording_duration_s=recording.duration_s,
            sample_window_ms=sample_window_ms,
            scanned_window_count=scan.scanned_window_count,
            rejected_clipped_windows=scan.rejected_clipped_windows,
            selection_note=scan.selection_note,
            sample_rate_hz=recording.sample_rate_hz,
            wsm_file=link.wsm_file,
            ws1_file=link.ws1_file,
            ws2_file=link.ws2_file,
        ))
    return tuple(representatives)


def analyze_typical_signals(
    database_path: Path,
    scene_type: str,
    polarization: str = "垂直极化",
    band: str = ALL_BAND,
    tolerance_mhz: float = 2.0,
    minimum_probability: float = 0.3,
) -> TypicalSignalResult:
    if tolerance_mhz <= 0:
        raise ValueError("频率匹配容差必须大于0 MHz")
    if not 0 <= minimum_probability <= 1:
        raise ValueError("最低出现概率必须位于0～1之间")

    locations = list_scene_locations(database_path)
    records = list_feature_records(database_path, polarization=polarization, band=band)
    record_map = {(record.city, record.point): record for record in records}
    valid_locations: dict[tuple[str, str], str] = {}
    observations: list[dict[str, object]] = []
    for location in locations:
        record = record_map.get((location.city, location.spectrum_point or location.point))
        if record is None:
            continue
        location_key = (location.city, location.point)
        valid_locations[location_key] = location.scene_type
        for peak in _safe_json_list(record.top_peaks_json):
            observations.append({
                "location_key": location_key,
                "scene_type": location.scene_type,
                "frequency": float(peak.get("frequency_mhz", 0.0)),
                "level": float(peak.get("field_dbuv_m", 0.0)),
                "bandwidth": float(peak.get("bandwidth_3db_mhz", 0.0)),
            })
    if not observations:
        raise ValueError("频谱特征库中没有可用于典型信号筛选的峰值")

    scene_locations = {key for key, value in valid_locations.items() if value == scene_type}
    if len(scene_locations) < 2:
        raise ValueError(f"{scene_type}至少需要两个具有频谱特征的地点")
    locations_by_scene: dict[str, set[tuple[str, str]]] = {}
    for key, value in valid_locations.items():
        locations_by_scene.setdefault(value, set()).add(key)

    clusters = _cluster_peak_observations(observations, tolerance_mhz)
    provisional: list[dict[str, object]] = []
    for cluster in clusters:
        target = [item for key, item in cluster.items() if key in scene_locations]
        if not target:
            continue
        scene_probability = len(target) / len(scene_locations)
        if scene_probability < minimum_probability:
            continue
        scene_probabilities: list[float] = []
        for other_scene, keys in locations_by_scene.items():
            if not keys:
                continue
            scene_probabilities.append(sum(key in cluster for key in keys) / len(keys))
        global_probability = float(np.mean(scene_probabilities)) if scene_probabilities else 0.0
        levels = np.asarray([float(item["level"]) for item in target])
        global_levels = np.asarray([float(item["level"]) for item in cluster.values()])
        bandwidths = np.asarray([float(item["bandwidth"]) for item in target])
        frequencies = np.asarray([float(item["frequency"]) for item in target])
        mean_level = float(np.mean(levels))
        global_median = float(np.median(global_levels))
        contrast = mean_level - global_median
        specificity = scene_probability - global_probability
        lift = scene_probability / max(global_probability, 1e-6)
        if global_probability >= 0.55 and specificity < 0.12 and contrast < 4.0:
            category = "公共背景"
        elif specificity >= 0.25 or lift >= 1.8:
            category = "场景特有"
        elif specificity >= 0.10 or contrast >= 4.0:
            category = "场景增强"
        else:
            category = "场景常见"
        provisional.append({
            "frequency": float(np.mean(frequencies)),
            "count": len(target),
            "scene_probability": scene_probability,
            "global_probability": global_probability,
            "specificity": specificity,
            "lift": lift,
            "mean_level": mean_level,
            "global_median": global_median,
            "contrast": contrast,
            "bandwidth": float(np.mean(bandwidths)),
            "category": category,
            "location_keys": tuple(sorted(
                (key for key in cluster if key in scene_locations), key=lambda item: (item[0].casefold(), item[1].casefold())
            )),
        })
    if not provisional:
        raise ValueError("当前阈值下没有筛选出典型信号")

    level_values = np.asarray([float(item["mean_level"]) for item in provisional])
    level_low, level_high = float(np.min(level_values)), float(np.max(level_values))
    for item in provisional:
        level_norm = (float(item["mean_level"]) - level_low) / max(level_high - level_low, 1e-9)
        item["score"] = (
            0.35 * float(item["scene_probability"])
            + 0.30 * max(float(item["specificity"]), 0.0)
            + 0.20 * min(max(float(item["contrast"]), 0.0) / 12.0, 1.0)
            + 0.15 * level_norm
        )
    provisional.sort(key=lambda item: (-float(item["score"]), -float(item["scene_probability"])))

    signals: list[TypicalSignal] = []
    for rank, item in enumerate(provisional, start=1):
        frequency = float(item["frequency"])
        location_keys = item["location_keys"]
        assert isinstance(location_keys, tuple)
        iq_candidates = _candidate_iq_links(database_path, location_keys, frequency)
        signals.append(TypicalSignal(
            rank=rank,
            scene_type=scene_type,
            typical_frequency_mhz=frequency,
            occurrence_count=int(item["count"]),
            scene_location_count=len(scene_locations),
            scene_probability=float(item["scene_probability"]),
            global_probability=float(item["global_probability"]),
            specificity=float(item["specificity"]),
            lift=float(item["lift"]),
            mean_level_dbuv_m=float(item["mean_level"]),
            global_median_level_dbuv_m=float(item["global_median"]),
            level_contrast_db=float(item["contrast"]),
            mean_bandwidth_3db_mhz=float(item["bandwidth"]),
            category=str(item["category"]),
            score=float(item["score"]),
            location_keys=location_keys,
            iq_candidates=iq_candidates,
        ))
    return TypicalSignalResult(
        scene_type=scene_type,
        polarization=polarization,
        band=band,
        tolerance_mhz=tolerance_mhz,
        minimum_probability=minimum_probability,
        scene_location_count=len(scene_locations),
        global_location_count=len(valid_locations),
        signals=tuple(signals),
    )


def _validate_waveform_request(sample_rate_hz: float, duration_s: float) -> int:
    if sample_rate_hz <= 0 or duration_s <= 0:
        raise ValueError("采样率和持续时间必须大于0")
    sample_count = int(round(sample_rate_hz * duration_s))
    if sample_count < 32:
        raise ValueError("重构波形采样点过少")
    if sample_count > MAX_RECONSTRUCTION_SAMPLES:
        raise ValueError(
            f"波形包含{sample_count:,}个复采样点，超过单次上限{MAX_RECONSTRUCTION_SAMPLES:,}。"
            "请缩短持续时间或降低采样率。"
        )
    return sample_count


def _bounded_waveform_request(
    sample_rate_hz: float,
    duration_s: float,
    maximum_buffer_samples: int = MAX_RECONSTRUCTION_SAMPLES,
) -> tuple[int, int, float]:
    """Return requested samples and a memory-safe loop-buffer size."""
    if sample_rate_hz <= 0 or duration_s <= 0:
        raise ValueError("采样率和持续时间必须大于0")
    requested_samples = int(round(sample_rate_hz * duration_s))
    if requested_samples < 32:
        raise ValueError("重构波形采样点过少")
    buffer_samples = min(requested_samples, maximum_buffer_samples)
    return requested_samples, buffer_samples, buffer_samples / sample_rate_hz


def _unit_peak(iq: np.ndarray, peak: float = 0.9) -> np.ndarray:
    iq = np.asarray(iq, dtype=np.complex64)
    maximum = float(np.max(np.abs(iq))) if iq.size else 0.0
    if maximum <= 1e-12:
        return iq
    return (iq * (peak / maximum)).astype(np.complex64)


def _unit_peak_inplace(iq: np.ndarray, peak: float = 0.9) -> np.ndarray:
    """Normalize a writable complex64 buffer without allocating another full IQ array."""
    maximum = float(np.max(np.abs(iq))) if iq.size else 0.0
    if maximum > 1e-12:
        iq *= peak / maximum
    return iq


def _resample_iq(iq: np.ndarray, source_rate_hz: float, target_rate_hz: float) -> np.ndarray:
    if math.isclose(source_rate_hz, target_rate_hz, rel_tol=1e-9):
        return np.asarray(iq, dtype=np.complex64)
    ratio = Fraction(target_rate_hz / source_rate_hz).limit_denominator(1000)
    return resample_poly(iq, ratio.numerator, ratio.denominator).astype(np.complex64)


def _crossfade_loop(segment: np.ndarray, target_samples: int, crossfade_samples: int) -> np.ndarray:
    segment = np.asarray(segment, dtype=np.complex64)
    if segment.size < 32:
        raise ValueError("实测片段过短，无法循环重构")
    if target_samples == segment.size:
        return segment
    if target_samples < segment.size:
        return segment[:target_samples].copy()
    crossfade_samples = min(max(0, crossfade_samples), segment.size // 3)
    if crossfade_samples == 0:
        return np.resize(segment, target_samples).astype(np.complex64)
    step = segment.size - crossfade_samples
    cycles = max(1, math.ceil((target_samples - segment.size) / step) + 1)
    output = np.zeros(segment.size + (cycles - 1) * step, dtype=np.complex64)
    weight = np.zeros(output.size, dtype=np.float32)
    fade_in = np.ones(segment.size, dtype=np.float32)
    fade_out = np.ones(segment.size, dtype=np.float32)
    ramp = (0.5 - 0.5 * np.cos(np.linspace(0.0, np.pi, crossfade_samples))).astype(np.float32)
    fade_in[:crossfade_samples] = ramp
    fade_out[-crossfade_samples:] = ramp[::-1]
    for cycle in range(cycles):
        start = cycle * step
        window = np.ones(segment.size, dtype=np.float32)
        if cycle > 0:
            window *= fade_in
        if cycle < cycles - 1:
            window *= fade_out
        output[start : start + segment.size] += segment * window
        weight[start : start + segment.size] += window
    output /= np.maximum(weight, 1e-6)
    return output[:target_samples]


def reconstruct_measured_signal(
    name: str,
    recording: IQRecording,
    start_s: float,
    source_duration_s: float,
    target_duration_s: float,
    target_sample_rate_hz: float | None = None,
    crossfade_s: float = 0.001,
    target_field_v_m: float = 10.0,
) -> ReconstructionResult:
    if start_s < 0 or source_duration_s <= 0:
        raise ValueError("实测片段起点不能为负，片段长度必须大于0")
    target_rate = float(target_sample_rate_hz or recording.sample_rate_hz)
    requested_samples, target_samples, buffer_duration_s = _bounded_waveform_request(
        target_rate,
        target_duration_s,
        MAX_MEASURED_PLAYBACK_SAMPLES,
    )
    source_start = int(round(start_s * recording.sample_rate_hz))
    source_count = int(round(source_duration_s * recording.sample_rate_hz))
    source_count = min(source_count, max(0, recording.total_samples - source_start))
    if source_count < 32:
        raise ValueError("所选实测片段超出记录范围")
    original_iq = read_iq_contiguous(recording, source_start, source_count)
    segment = np.array(original_iq, dtype=np.complex64, copy=True)
    segment -= np.mean(segment)
    segment = _resample_iq(segment, recording.sample_rate_hz, target_rate)
    segment = _unit_peak_inplace(segment, 0.8)
    source_looped = target_samples > segment.size
    buffer_looped = requested_samples > target_samples
    iq = _crossfade_loop(segment, target_samples, int(round(crossfade_s * target_rate)))
    del segment
    iq = _unit_peak_inplace(iq)
    actual_source_duration_s = source_count / recording.sample_rate_hz
    component = ReconstructionComponent(
        name=recording.stem,
        source_type="实测IQ",
        modulation="原始采集波形",
        frequency_mhz=recording.center_frequency_mhz,
        offset_mhz=0.0,
        relative_level_db=0.0,
        bandwidth_mhz=recording.sample_rate_hz / 1e6,
        source_reference=f"{recording.stem}@{start_s:.6f}～{start_s + actual_source_duration_s:.6f}s",
    )
    metadata = {
        "source_recording": recording.stem,
        "source_reference_level_dbm": recording.reference_level_dbm,
        "source_start_s": start_s,
        "source_duration_s": actual_source_duration_s,
        "source_end_s": start_s + actual_source_duration_s,
        "crossfade_s": crossfade_s,
        "requested_playback_duration_s": target_duration_s,
        "waveform_buffer_duration_s": buffer_duration_s,
        "source_segment_looped": source_looped,
        "device_buffer_looped": buffer_looped,
        "loop_playback_required": source_looped or buffer_looped,
        "target_field_v_m": target_field_v_m,
        "absolute_field_note": "实采IQ保持波形特征；目标V/m需在实验室通过信号源、功放、天线和场强探头闭环标定。",
        "processing": [
            "去除复均值",
            "重采样" if target_rate != recording.sample_rate_hz else "保持原采样率",
            "交叉渐变循环" if source_looped else "保留完整入选片段",
            "峰值归一化",
        ],
    }
    return _finish_result(
        name,
        "实际采集典型场景信号重构",
        iq,
        target_rate,
        recording.center_frequency_mhz,
        (component,),
        metadata,
        original_iq=original_iq,
        original_sample_rate_hz=recording.sample_rate_hz,
        original_center_frequency_mhz=recording.center_frequency_mhz,
    )


def generate_modulated_signal(
    modulation: str,
    sample_rate_hz: float,
    duration_s: float,
    offset_hz: float = 0.0,
    amplitude: float = 1.0,
    modulation_frequency_hz: float = 1_000.0,
    modulation_index: float = 0.5,
    symbol_rate: float = 10_000.0,
    seed: int = 2026,
) -> np.ndarray:
    count = _validate_waveform_request(sample_rate_hz, duration_s)
    t = np.arange(count, dtype=np.float64) / sample_rate_hz
    carrier_phase = 2.0 * np.pi * offset_hz * t
    mode = modulation.upper().strip()
    rng = np.random.default_rng(seed)
    if mode in {"CW", "单载波"}:
        iq = np.exp(1j * carrier_phase)
    elif mode == "AM":
        depth = min(max(modulation_index, 0.0), 0.99)
        envelope = 1.0 + depth * np.sin(2.0 * np.pi * modulation_frequency_hz * t)
        iq = envelope * np.exp(1j * carrier_phase)
    elif mode == "FM":
        beta = max(modulation_index, 0.0)
        phase = carrier_phase + beta * np.sin(2.0 * np.pi * modulation_frequency_hz * t)
        iq = np.exp(1j * phase)
    elif mode in {"ASK", "FSK", "BPSK", "QPSK", "16QAM", "QAM"}:
        samples_per_symbol = max(2, int(round(sample_rate_hz / max(symbol_rate, 1.0))))
        symbol_count = math.ceil(count / samples_per_symbol)
        if mode == "ASK":
            depth = min(max(modulation_index, 0.0), 1.0)
            symbols = rng.choice(np.asarray([1.0 - depth, 1.0], dtype=np.float32), symbol_count).astype(np.complex64)
        elif mode == "FSK":
            symbols = rng.choice(np.asarray([-1.0, 1.0], dtype=np.float32), symbol_count)
            deviation_hz = max(symbol_rate * max(modulation_index, 0.1) / 2.0, 1.0)
            instantaneous = np.repeat(symbols, samples_per_symbol)[:count] * deviation_hz
            phase = carrier_phase + 2.0 * np.pi * np.cumsum(instantaneous) / sample_rate_hz
            iq = np.exp(1j * phase)
            return (amplitude * np.asarray(iq, dtype=np.complex64)).astype(np.complex64)
        elif mode == "BPSK":
            symbols = rng.choice(np.asarray([-1.0, 1.0], dtype=np.float32), symbol_count).astype(np.complex64)
        elif mode == "QPSK":
            levels = rng.choice(np.asarray([-1.0, 1.0], dtype=np.float32), (2, symbol_count))
            symbols = ((levels[0] + 1j * levels[1]) / math.sqrt(2.0)).astype(np.complex64)
        else:
            levels = rng.choice(np.asarray([-3.0, -1.0, 1.0, 3.0], dtype=np.float32), (2, symbol_count))
            symbols = ((levels[0] + 1j * levels[1]) / math.sqrt(10.0)).astype(np.complex64)
        baseband = np.repeat(symbols, samples_per_symbol)[:count]
        cutoff = min(0.45 * sample_rate_hz, max(symbol_rate * 0.7, sample_rate_hz / count))
        taps = firwin(129, cutoff=cutoff, fs=sample_rate_hz)
        baseband = lfilter(taps, [1.0], baseband).astype(np.complex64)
        iq = baseband * np.exp(1j * carrier_phase)
    elif mode in {"PULSE", "脉冲"}:
        frequency = max(modulation_frequency_hz, 1.0)
        duty = min(max(modulation_index, 0.01), 0.95)
        gate = (np.mod(t, 1.0 / frequency) < duty / frequency).astype(np.float32)
        iq = gate * np.exp(1j * carrier_phase)
    elif mode in {"NOISE", "噪声"}:
        iq = rng.normal(size=count) + 1j * rng.normal(size=count)
    else:
        raise ValueError(f"不支持的调制类型：{modulation}")
    return (amplitude * np.asarray(iq, dtype=np.complex64)).astype(np.complex64)


def reconstruct_single_modulated(
    name: str,
    modulation: str,
    center_frequency_mhz: float,
    sample_rate_hz: float,
    duration_s: float,
    offset_mhz: float = 0.0,
    relative_level_db: float = 0.0,
    modulation_frequency_hz: float = 1_000.0,
    modulation_index: float = 0.5,
    symbol_rate: float = 10_000.0,
    seed: int = 2026,
    target_field_v_m: float = 30.0,
) -> ReconstructionResult:
    if abs(offset_mhz) * 1e6 >= sample_rate_hz / 2:
        raise ValueError("频偏超出输出采样带宽")
    requested_samples, buffer_samples, buffer_duration_s = _bounded_waveform_request(sample_rate_hz, duration_s)
    amplitude = 10.0 ** (relative_level_db / 20.0)
    iq = generate_modulated_signal(
        modulation, sample_rate_hz, buffer_duration_s, offset_mhz * 1e6, amplitude,
        modulation_frequency_hz, modulation_index, symbol_rate, seed,
    )
    iq = _unit_peak(iq)
    bandwidth = symbol_rate / 1e6 if modulation.upper() in {"ASK", "FSK", "BPSK", "QPSK", "16QAM", "QAM"} else 0.0
    component = ReconstructionComponent(
        name=name,
        source_type="参数化合成",
        modulation=modulation,
        frequency_mhz=center_frequency_mhz + offset_mhz,
        offset_mhz=offset_mhz,
        relative_level_db=relative_level_db,
        bandwidth_mhz=bandwidth,
    )
    metadata = {
        "modulation_frequency_hz": modulation_frequency_hz,
        "modulation_index": modulation_index,
        "symbol_rate": symbol_rate,
        "seed": seed,
        "requested_playback_duration_s": duration_s,
        "waveform_buffer_duration_s": buffer_duration_s,
        "loop_playback_required": requested_samples > buffer_samples,
        "target_field_v_m": target_field_v_m,
        "absolute_field_note": "目标场强用于回放配置；数字IQ到V/m的映射需通过信号源、功放和天线链路标定。",
    }
    return _finish_result(name, "单频点调制信号重构", iq, sample_rate_hz, center_frequency_mhz, (component,), metadata)


def parse_component_lines(text: str) -> list[dict[str, object]]:
    components: list[dict[str, object]] = []
    for line_number, raw_line in enumerate(text.splitlines(), start=1):
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        values = [value.strip() for value in line.split(",")]
        if len(values) < 5:
            raise ValueError(f"第{line_number}行需要5列：名称,调制,频偏MHz,相对电平dB,调制频率或符号率")
        name, modulation = values[:2]
        try:
            offset_mhz, level_db, parameter = map(float, values[2:5])
        except ValueError as exc:
            raise ValueError(f"第{line_number}行存在非数字参数") from exc
        components.append({
            "name": name,
            "modulation": modulation,
            "offset_mhz": offset_mhz,
            "level_db": level_db,
            "parameter": parameter,
        })
    if not components:
        raise ValueError("没有可合成的无线电系统信号")
    return components


def reconstruct_multi_system(
    name: str,
    center_frequency_mhz: float,
    sample_rate_hz: float,
    duration_s: float,
    component_lines: str,
    seed: int = 2026,
    target_field_v_m: float = 30.0,
) -> ReconstructionResult:
    specifications = parse_component_lines(component_lines)
    requested_samples, buffer_samples, buffer_duration_s = _bounded_waveform_request(sample_rate_hz, duration_s)
    total = np.zeros(buffer_samples, dtype=np.complex64)
    components: list[ReconstructionComponent] = []
    for index, specification in enumerate(specifications):
        modulation = str(specification["modulation"])
        offset_mhz = float(specification["offset_mhz"])
        if abs(offset_mhz) * 1e6 >= sample_rate_hz / 2:
            raise ValueError(f"{specification['name']}的频偏超出采样带宽")
        parameter = float(specification["parameter"])
        is_digital = modulation.upper() in {"ASK", "FSK", "BPSK", "QPSK", "16QAM", "QAM"}
        signal = generate_modulated_signal(
            modulation=modulation,
            sample_rate_hz=sample_rate_hz,
            duration_s=buffer_duration_s,
            offset_hz=offset_mhz * 1e6,
            amplitude=10.0 ** (float(specification["level_db"]) / 20.0),
            modulation_frequency_hz=parameter if not is_digital else 1_000.0,
            symbol_rate=parameter if is_digital else 10_000.0,
            seed=seed + index,
        )
        total += signal
        components.append(ReconstructionComponent(
            name=str(specification["name"]),
            source_type="参数化合成",
            modulation=modulation,
            frequency_mhz=center_frequency_mhz + offset_mhz,
            offset_mhz=offset_mhz,
            relative_level_db=float(specification["level_db"]),
            bandwidth_mhz=parameter / 1e6 if is_digital else 0.0,
        ))
    total = _unit_peak(total)
    return _finish_result(
        name, "多无线电系统合成信号重构", total, sample_rate_hz, center_frequency_mhz,
        tuple(components), {
            "seed": seed, "component_definition": component_lines, "target_field_v_m": target_field_v_m,
            "requested_playback_duration_s": duration_s,
            "waveform_buffer_duration_s": buffer_duration_s,
            "loop_playback_required": requested_samples > buffer_samples,
            "absolute_field_note": "目标场强用于回放配置；各系统相对电平由数字波形保持，绝对V/m需实验室链路标定。",
        },
    )


def reconstruct_hybrid_scene(
    name: str,
    recording: IQRecording,
    source_start_s: float,
    source_duration_s: float,
    target_duration_s: float,
    target_sample_rate_hz: float,
    crossfade_s: float,
    component_lines: str,
    target_field_v_m: float = 30.0,
    seed: int = 2026,
) -> ReconstructionResult:
    """Combine one measured representative IQ background with simulated in-band systems."""
    requested_samples, target_samples, buffer_duration_s = _bounded_waveform_request(
        target_sample_rate_hz, target_duration_s
    )
    source_start = int(round(source_start_s * recording.sample_rate_hz))
    source_count = min(
        int(round(source_duration_s * recording.sample_rate_hz)),
        max(0, recording.total_samples - source_start),
    )
    if source_count < 32:
        raise ValueError("复杂场景的实测背景片段超出记录范围")
    original_iq = read_iq_contiguous(recording, source_start, source_count)
    measured = (original_iq - np.mean(original_iq)).astype(np.complex64)
    measured = _resample_iq(measured, recording.sample_rate_hz, target_sample_rate_hz)
    measured = _crossfade_loop(
        _unit_peak(measured, 0.65), target_samples, int(round(crossfade_s * target_sample_rate_hz))
    )
    total = measured.astype(np.complex64)
    components = [ReconstructionComponent(
        name=recording.stem,
        source_type="实测典型场景IQ",
        modulation="原始采集波形",
        frequency_mhz=recording.center_frequency_mhz,
        offset_mhz=0.0,
        relative_level_db=0.0,
        bandwidth_mhz=recording.sample_rate_hz / 1e6,
        source_reference=f"{recording.stem}@{source_start_s:.6f}s",
    )]
    specifications = parse_component_lines(component_lines)
    for index, specification in enumerate(specifications):
        modulation = str(specification["modulation"])
        offset_mhz = float(specification["offset_mhz"])
        if abs(offset_mhz) * 1e6 >= target_sample_rate_hz / 2:
            raise ValueError(f"{specification['name']}的频偏超出当前回放通道带宽")
        parameter = float(specification["parameter"])
        is_digital = modulation.upper() in {"ASK", "FSK", "BPSK", "QPSK", "16QAM", "QAM"}
        level_db = float(specification["level_db"])
        generated = generate_modulated_signal(
            modulation=modulation,
            sample_rate_hz=target_sample_rate_hz,
            duration_s=buffer_duration_s,
            offset_hz=offset_mhz * 1e6,
            amplitude=10.0 ** (level_db / 20.0),
            modulation_frequency_hz=parameter if not is_digital else 1_000.0,
            symbol_rate=parameter if is_digital else 10_000.0,
            seed=seed + index,
        )
        total += generated
        components.append(ReconstructionComponent(
            name=str(specification["name"]), source_type="参数化注入", modulation=modulation,
            frequency_mhz=recording.center_frequency_mhz + offset_mhz, offset_mhz=offset_mhz,
            relative_level_db=level_db, bandwidth_mhz=parameter / 1e6 if is_digital else 0.0,
        ))
    total = _unit_peak(total)
    metadata = {
        "scene_structure": "实采典型场景背景 + 同通道模拟无线电/干扰信号",
        "source_recording": recording.stem,
        "source_start_s": source_start_s,
        "source_duration_s": source_duration_s,
        "component_definition": component_lines,
        "requested_playback_duration_s": target_duration_s,
        "waveform_buffer_duration_s": buffer_duration_s,
        "loop_playback_required": requested_samples > target_samples,
        "target_field_v_m": target_field_v_m,
        "seed": seed,
        "absolute_field_note": "单通道数字波形保持相对幅度；多频段复杂场景应分通道生成并在实验室同步合路，绝对V/m需标定。",
    }
    return _finish_result(
        name, "复杂场景多信号组合重构", total, target_sample_rate_hz,
        recording.center_frequency_mhz, tuple(components), metadata,
        original_iq=original_iq,
        original_sample_rate_hz=recording.sample_rate_hz,
        original_center_frequency_mhz=recording.center_frequency_mhz,
    )


def _extract_measured_channel(
    candidate: IQCandidate,
    signal_frequency_mhz: float,
    bandwidth_mhz: float,
    target_rate_hz: float,
    target_samples: int,
) -> np.ndarray:
    recording = candidate.recording()
    source_samples = min(recording.total_samples, max(32_768, int(round(0.02 * recording.sample_rate_hz))))
    iq = read_iq_contiguous(recording, 0, source_samples)
    iq = (iq - np.mean(iq)).astype(np.complex64)
    source_offset_hz = (signal_frequency_mhz - recording.center_frequency_mhz) * 1e6
    time = np.arange(iq.size, dtype=np.float64) / recording.sample_rate_hz
    baseband = iq * np.exp(-1j * 2.0 * np.pi * source_offset_hz * time)
    cutoff_hz = min(
        recording.sample_rate_hz * 0.42,
        max(50_000.0, bandwidth_mhz * 0.75e6),
    )
    taps = firwin(257, cutoff=cutoff_hz, fs=recording.sample_rate_hz)
    baseband = lfilter(taps, [1.0], baseband).astype(np.complex64)
    baseband = _resample_iq(baseband, recording.sample_rate_hz, target_rate_hz)
    baseband = _unit_peak(baseband, 1.0)
    return _crossfade_loop(baseband, target_samples, min(baseband.size // 4, int(0.001 * target_rate_hz)))


def reconstruct_complex_scene(
    name: str,
    signals: Iterable[TypicalSignal],
    center_frequency_mhz: float,
    sample_rate_hz: float,
    duration_s: float,
    measured_first: bool = True,
    seed: int = 2026,
) -> ReconstructionResult:
    selected = list(signals)
    if not selected:
        raise ValueError("请至少选择一个典型信号")
    target_samples = _validate_waveform_request(sample_rate_hz, duration_s)
    half_band_mhz = sample_rate_hz / 2e6
    for signal in selected:
        guard = max(signal.mean_bandwidth_3db_mhz / 2.0, 0.05)
        if abs(signal.typical_frequency_mhz - center_frequency_mhz) + guard >= half_band_mhz:
            raise ValueError(
                f"{signal.typical_frequency_mhz:.6f} MHz超出当前{sample_rate_hz / 1e6:.3f} MS/s输出带宽。"
                "请调整重构中心频率、提高采样率或分频段生成。"
            )
    reference_level = max(signal.mean_level_dbuv_m for signal in selected)
    rng = np.random.default_rng(seed)
    total = np.zeros(target_samples, dtype=np.complex64)
    components: list[ReconstructionComponent] = []
    for index, signal in enumerate(selected):
        offset_mhz = signal.typical_frequency_mhz - center_frequency_mhz
        relative_level = signal.mean_level_dbuv_m - reference_level
        amplitude = 10.0 ** (relative_level / 20.0)
        source_type = "参数化合成"
        source_reference = ""
        modulation = "QPSK" if signal.mean_bandwidth_3db_mhz >= 0.3 else "FM"
        if measured_first and signal.iq_candidates:
            try:
                candidate = signal.iq_candidates[0]
                baseband = _extract_measured_channel(
                    candidate, signal.typical_frequency_mhz, signal.mean_bandwidth_3db_mhz,
                    sample_rate_hz, target_samples,
                )
                time = np.arange(target_samples, dtype=np.float64) / sample_rate_hz
                component_iq = baseband * np.exp(1j * 2.0 * np.pi * offset_mhz * 1e6 * time)
                source_type = "实测IQ"
                source_reference = f"{candidate.city}/{candidate.point}/{candidate.recording_stem}"
                modulation = "实测片段"
            except (FileNotFoundError, OSError, ValueError):
                component_iq = generate_modulated_signal(
                    modulation, sample_rate_hz, duration_s, offset_mhz * 1e6,
                    symbol_rate=max(10_000.0, signal.mean_bandwidth_3db_mhz * 0.6e6), seed=seed + index,
                )
        else:
            component_iq = generate_modulated_signal(
                modulation, sample_rate_hz, duration_s, offset_mhz * 1e6,
                modulation_frequency_hz=max(200.0, min(15_000.0, signal.mean_bandwidth_3db_mhz * 0.1e6)),
                modulation_index=2.0 if modulation == "FM" else 0.5,
                symbol_rate=max(10_000.0, signal.mean_bandwidth_3db_mhz * 0.6e6),
                seed=seed + index,
            )
        if signal.category in {"场景特有", "场景增强"}:
            period = rng.uniform(0.008, 0.025)
            duty = min(max(signal.scene_probability, 0.2), 0.9)
            time = np.arange(target_samples, dtype=np.float64) / sample_rate_hz
            phase = rng.uniform(0.0, period)
            gate = (np.mod(time + phase, period) < period * duty).astype(np.float32)
            component_iq *= gate
        total += (amplitude * component_iq).astype(np.complex64)
        components.append(ReconstructionComponent(
            name=f"{signal.category}-{signal.typical_frequency_mhz:.6f}MHz",
            source_type=source_type,
            modulation=modulation,
            frequency_mhz=signal.typical_frequency_mhz,
            offset_mhz=offset_mhz,
            relative_level_db=relative_level,
            bandwidth_mhz=signal.mean_bandwidth_3db_mhz,
            source_reference=source_reference,
        ))
    total = _unit_peak(total)
    metadata = {
        "scene_type": selected[0].scene_type,
        "measured_first": measured_first,
        "seed": seed,
        "signal_categories": {category: sum(item.category == category for item in selected) for category in {item.category for item in selected}},
        "absolute_field_note": "数字幅度保持相对场强关系；绝对V/m需在实验室回放链路中标定。",
    }
    return _finish_result(
        name, "复杂场景多信号组合重构", total, sample_rate_hz, center_frequency_mhz,
        tuple(components), metadata,
    )


def _make_preview_figure(
    name: str,
    iq: np.ndarray,
    sample_rate_hz: float,
    center_frequency_mhz: float,
    components: tuple[ReconstructionComponent, ...],
) -> Figure:
    figure = Figure(figsize=(10.5, 7.2), constrained_layout=True)
    time_ax, spectrum_ax, spectrogram_ax = figure.subplots(3, 1)
    iq_window_samples = min(iq.size, max(32, int(round(0.001 * sample_rate_hz))))
    iq_stride = max(1, math.ceil(iq_window_samples / 5000))
    iq_indices = np.arange(0, iq_window_samples, iq_stride)
    time_ms = iq_indices / sample_rate_hz * 1e3
    time_ax.plot(time_ms, iq[iq_indices].real, linewidth=0.7, color="#2563eb", label="I")
    time_ax.plot(time_ms, iq[iq_indices].imag, linewidth=0.7, color="#dc2626", alpha=0.82, label="Q")
    time_ax.set(title="重构信号I/Q波形（1 ms窗口）", xlabel="窗口内时间 (ms)", ylabel="归一化幅度")
    time_ax.legend(loc="upper right", ncol=2)
    time_ax.grid(alpha=0.22)

    fft_points = min(iq.size, 262_144)
    fft_iq = iq[:fft_points]
    window = np.hanning(fft_points)
    spectrum = np.fft.fftshift(np.fft.fft(fft_iq * window))
    psd = 20.0 * np.log10(np.maximum(np.abs(spectrum), 1e-12))
    psd -= float(np.max(psd))
    frequencies = center_frequency_mhz + np.fft.fftshift(np.fft.fftfreq(fft_points, 1.0 / sample_rate_hz)) / 1e6
    spectrum_ax.plot(frequencies, psd, linewidth=0.8, color="#0f766e")
    spectrum_ax.set(title="重构信号相对频谱", xlabel="频率 (MHz)", ylabel="相对幅度 (dB)")
    spectrum_ax.set_ylim(-100, 3)
    spectrum_ax.grid(alpha=0.22)
    for component in components:
        spectrum_ax.axvline(component.frequency_mhz, color="#ea580c", alpha=0.45, linewidth=0.8)

    nfft = min(2048, 2 ** int(math.floor(math.log2(iq.size))))
    available_windows = max(1, iq.size - nfft + 1)
    slice_count = min(600, max(1, math.ceil(iq.size / nfft)))
    starts = np.linspace(0, available_windows - 1, slice_count, dtype=np.int64)
    spec_window = np.hanning(nfft).astype(np.float32)
    spec_matrix = np.empty((nfft, slice_count), dtype=np.float32)
    for column, start in enumerate(starts):
        window_iq = iq[start : start + nfft]
        transformed = np.fft.fftshift(np.fft.fft(window_iq * spec_window))
        spec_matrix[:, column] = 20.0 * np.log10(np.maximum(np.abs(transformed), 1e-12))
    spec_matrix -= float(np.max(spec_matrix))
    spec_time_ms = (starts + nfft / 2.0) / sample_rate_hz * 1e3
    baseband_frequency_mhz = np.fft.fftshift(np.fft.fftfreq(nfft, 1.0 / sample_rate_hz)) / 1e6
    spectrogram_ax.imshow(
        spec_matrix,
        origin="lower",
        aspect="auto",
        extent=(
            float(spec_time_ms[0]), float(spec_time_ms[-1]),
            float(baseband_frequency_mhz[0]), float(baseband_frequency_mhz[-1]),
        ),
        vmin=-80,
        vmax=0,
        cmap="viridis",
        interpolation="nearest",
    )
    spectrogram_ax.set(title="重构信号全时长时频图", xlabel="重构时间 (ms)", ylabel="基带频率偏移 (MHz)")
    figure.suptitle(f"{name}｜中心频率 {center_frequency_mhz:.6f} MHz｜{sample_rate_hz / 1e6:.3f} MS/s")
    return figure


def _finish_result(
    name: str,
    mode: str,
    iq: np.ndarray,
    sample_rate_hz: float,
    center_frequency_mhz: float,
    components: tuple[ReconstructionComponent, ...],
    metadata: dict[str, object],
    original_iq: np.ndarray | None = None,
    original_sample_rate_hz: float | None = None,
    original_center_frequency_mhz: float | None = None,
) -> ReconstructionResult:
    iq = np.asarray(iq, dtype=np.complex64)
    figure = _make_preview_figure(name, iq, sample_rate_hz, center_frequency_mhz, components)
    return ReconstructionResult(
        name,
        mode,
        iq,
        sample_rate_hz,
        center_frequency_mhz,
        components,
        metadata,
        figure,
        None if original_iq is None else np.asarray(original_iq, dtype=np.complex64),
        original_sample_rate_hz,
        original_center_frequency_mhz,
    )


def reconstruction_explanation(mode: str) -> str:
    explanations = {
        "实际采集典型场景信号重构": (
            "先依据多地点IQ覆盖和频谱证据确定代表频段，再用短功率窗定位候选IQ中的最强稳定时刻，"
            "围绕该时刻截取较长的连续原始IQ；片段短于目标时长时采用交叉渐变循环，生成可连续回放的IQ。"
        ),
        "单频点调制信号重构": (
            "根据调制、频偏、符号率和电平等参数生成新的数字IQ，因此没有一一对应的原始采集信号。"
        ),
        "多无线电系统合成信号重构": (
            "按参数分别生成多个无线电分量并在同一IQ通道中叠加，因此没有单一的原始采集信号。"
        ),
        "复杂场景多信号组合重构": (
            "以原始采集IQ作为场景背景，再叠加参数化无线电或干扰分量，形成可回放的复杂场景IQ。"
        ),
    }
    return explanations.get(mode, "将选定或生成的信号处理为可由设备连续回放的数字IQ波形。")


def reconstruction_summary(result: ReconstructionResult) -> str:
    rms = float(np.sqrt(np.mean(np.abs(result.iq) ** 2)))
    peak = float(np.max(np.abs(result.iq)))
    requested_duration_s = float(result.metadata.get("requested_playback_duration_s", result.duration_s))
    loop_required = bool(result.metadata.get("loop_playback_required", False))
    source_duration_s = result.metadata.get("source_duration_s")
    source_start_s = result.metadata.get("source_start_s")
    lines = [
        f"名称：{result.name}",
        f"方式：{result.mode}",
        f"重构含义：{reconstruction_explanation(result.mode)}",
        f"中心频率：{result.center_frequency_mhz:.6f} MHz",
        f"采样率：{result.sample_rate_hz / 1e6:.6f} MS/s",
        f"目标回放时长：{requested_duration_s * 1e3:.3f} ms",
        f"波形缓冲区时长：{result.duration_s * 1e3:.3f} ms"
        f"（{'循环播放至目标时长' if loop_required else '无需循环扩展'}）",
        (
            f"入选实测片段：{float(source_start_s):.6f}～"
            f"{float(source_start_s) + float(source_duration_s):.6f} s"
            f"（{float(source_duration_s) * 1e3:.3f} ms）"
            if source_duration_s is not None and source_start_s is not None
            else "入选实测片段：不适用"
        ),
        f"复采样点：{result.iq.size:,}",
        f"RMS / 峰值：{rms:.6f} / {peak:.6f}",
        f"组成信号：{len(result.components)} 个",
        f"目标场强：{float(result.metadata['target_field_v_m']):g} V/m" if "target_field_v_m" in result.metadata else "目标场强：未设置",
        "",
    ]
    for index, component in enumerate(result.components, start=1):
        lines.append(
            f"{index}. {component.name}｜{component.source_type}｜{component.modulation}｜"
            f"{component.frequency_mhz:.6f} MHz｜{component.relative_level_db:.2f} dB"
        )
        if component.source_reference:
            lines.append(f"   来源：{component.source_reference}")
    lines.extend(("", "说明：导出的数字IQ仅保持相对幅度；绝对场强需在实验室回放链路中标定。"))
    return "\n".join(lines)


def save_reconstruction(result: ReconstructionResult, output_root: Path) -> tuple[Path, ...]:
    safe_name = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff._-]+", "_", result.name).strip("._") or "reconstruction"
    output_dir = output_root / "reconstructed_signals" / safe_name
    output_dir.mkdir(parents=True, exist_ok=True)
    npy_path = output_dir / f"{safe_name}.npy"
    binary_path = output_dir / f"{safe_name}_float32_iq.bin"
    json_path = output_dir / f"{safe_name}.json"
    figure_path = output_dir / f"{safe_name}_preview.png"
    np.save(npy_path, result.iq.astype(np.complex64), allow_pickle=False)
    interleaved = np.empty(result.iq.size * 2, dtype="<f4")
    interleaved[0::2] = result.iq.real
    interleaved[1::2] = result.iq.imag
    interleaved.tofile(binary_path)
    payload = {
        "name": result.name,
        "mode": result.mode,
        "sample_rate_hz": result.sample_rate_hz,
        "center_frequency_mhz": result.center_frequency_mhz,
        "duration_s": result.duration_s,
        "requested_playback_duration_s": float(
            result.metadata.get("requested_playback_duration_s", result.duration_s)
        ),
        "loop_playback_required": bool(result.metadata.get("loop_playback_required", False)),
        "sample_count": int(result.iq.size),
        "binary_format": "little-endian float32 interleaved I,Q",
        "components": [asdict(component) for component in result.components],
        "metadata": result.metadata,
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    result.figure.savefig(figure_path, dpi=150)
    return npy_path, binary_path, json_path, figure_path
